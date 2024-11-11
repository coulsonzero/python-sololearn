"""
创建工作簿：
from openpyxl import workbook
wb = workbook()
ws = wb.active
创建新的工作表：wb.creat_sheet()
ws1 = wb.create_sheet("Mysheet")
更改表名：
worksheet.title
ws.title = "New Title"

改色：Worksheet.sheet_properties.tabColor ="1072BA"
查看所有工作表名称：print(wb.sheetnames)
['Sheet2', 'New Title', 'Sheet1']
循环浏览工作表：
for sheet in wb:
...print(sheet.title)
在单个工作簿中创建工作表副本：wb.copy_worksheet()
source = wb.active
target = wb.copy_worksheet(source)


读取文件： openpyxl.load_workbook()
保存文件：
wb = workbook()
wb.save(' .xlsx')

# 访问一个单元格:
c = ws['A4']/Worksheet.cell()
# 访问多个单元格：
cell_range = ws['A1': 'C2']
row=x
columb=y
cell

# 写数据：
sheet['C3'] = " "

# 访问单元格数据：
for cell in sheet["B2:B4"]:
    print(cell[0],value)

for row in sheet.iter_rows(min_row=2,max_row=5,max_col=6): #指定行列
    for cell in row:
        print(cell.value,end=",")#,or空格
        print()

"""


"""
1.IDE: PyCharm
2.安装openpyxl库方法：
* Pycharm —— File —— Settings —— Project —— Python interpreter"+" —— openpyxl
* win+R —— cmd —— pip install openpyxl
3.网页查阅“openpyxl”使用说明
4.Python文档： "docs.python.org"


Excel文件：workbook
当前工作表：sheet/worksheet/ws
单元格：cell
行：row
列：columb


一、Excel文件操作：
* 读取文件: openpyxl.load_Workbook()
* 创建空白文件：wb=workbook()
* 保存文件：wb.save('filename.xlsx')
* 查看文件名：print(wb)

创建空白工作簿：
from openpyxl import workbook
wb = workbook()
ws = wb.active       #ws=sheet

读取Excel文件和更换当前工作表
import openpyxl as xl
wb = xl.load_workbook("filename.xlsx") 
ws = wb.active            
ws = wb["Sheet1"]                       


#导入openpyxl库
import openpyxl as xl

#打开excel文件
def process_workbook(filename):     
    wb = xl.load_workbook(filename)  
   #wb = xl.load_workbook(filename=' ', read_only=True) 当写入大文件时插入只读，将使用延迟加载
#获取和更换工作表单
    ws = wb.active
    ws = wb["Sheet1"]    #ws=sheet
#保存文件
    wb.save(filename)
    
    
二、工作表操作方法：
* 查看当前工作表：
ws = wb.active
print（ws）   
* （并排遍历）查看所有工作表名：print(wb.sheetnames)
      （并列遍历）查看所有工作表名：   
for sheet in wb:
    print(sheet.title)
    
* 替换当前工作表：ws = wb["Sheet1"] 
* 更改表名：表名.title = "New Title"
* 创建新的工作表：new_sheet = wb.create_sheet("表名"）
* 创建工作表副本：            wb.copy_worksheet("表名")
* 改色：sheet.sheet_properties.tabColor =  "1072BA"


三、单元格
#单元格赋值
ws['C3'] = " "         
ws.cell（row,col,value)

#访问某一个单元格数据
cell = ws["A2"]  #cell = ws.cell(row,col)
print(cell.value)

#访问特定行列单元格数据
for row in ws.iter_rows(min_row=2,max_row=5,max_col=6):  
    for cell in row:
        print(cell.value,end=",")    #,or多个空格
    print()

#访问所有单元格数据
for row in ws:
    for cell in row:
        print(cell.value,end="  ")
    print()



#合并单元格
ws.merge_cells('A2:D2')

ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
#取消合并单元格
ws.unmerge_cells('A2:D2')

添加图片
pillow库  #但无需import
#添加图片及信息
ws.merge_cells('A7:E7')
ws['A7'] = 'You could see the picture below'
img = Image('刘备.jpg')
ws.add_image(img, 'A8')


创建图表及类型
'''
面积图：AreaChart
3D面积图：AreaChart3D
垂直条形图：BarChart
3D垂直条形图：BarChart3D
重叠条形图：StackedChart
折线图：LineChart
散点图：ScatterChart
饼状图：PieChart
'''


import openpyxl
from openpyxl.chart import *Chart, Reference

def process_workbook(filename):
    ...
    values = Reference(ws,
                   min_row= ,
                   max_row= ,
                   min_col= ,
                   max_col= )

    chart = *Chart()
    chart.add_data(values)
    ws.add_chart(chart,'D3')
    
    wb.save(filename)


import openpyxl as xl
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference


#读取Excel文件
wb = xl.load_workbook("Excel 1.xlsx")
#设置活跃工作表
ws = wb.active
#查看活跃工作表
print(ws)
#创建新的工作表
new_sheet = wb.create_sheet("Newsheet")
#（并排）查看所有工作表
print(wb.sheetnames)
#更换表名
new_sheet.title = "sheet2"
print(wb.sheetnames)


#查看某个单元格数据
cell = ws["D5"]
print(cell.value)
cell = ws.cell(3,4)
print(cell.value)
#单元格赋值
ws['E3']='120'
print(ws['E3'].value)


ws.cell(4,5,150)
print(ws.cell(4,5,150).value)
#合并单元格
ws.merge_cells('A1:E1')
ws.cell(1,1,'三国演义')


#添加图片及信息
ws.merge_cells('A7:E7')
ws['A7'] = 'You could see the picture below'
img = Image('刘备.jpg')
ws.add_image(img, 'A8')


#遍历特定行列
for col in ws.iter_rows(min_col=2,max_col=6,max_row=4):  #指定行列
    for cell in col:
        print(cell.value,end="         ")    #,or空格
    print()
#遍历所有行列
for row in ws:
    for cell in row:
        print(cell.value,end="  ")
    print()

#添加图表
values = Reference(ws,
                   min_row=3,
                   max_row=6,
                   min_col=4,
                   max_col=4)


chart1 = BarChart()
chart1.add_data(values)
ws.add_chart(chart1, 'F2')
chart1 = BarChart()

#保存文件数据
wb.save("New Excel.xlsx")
"""